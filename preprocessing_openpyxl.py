import os
import openpyxl
from pathlib import Path

import config
from utility_functions import terminate_script, catch_errors

'''
Предварительная обработка файлов Excel:
    - снятие объединение ячеек
    - удаление пустых столбцов
    - нумерация уровней вложенности (отдельный столбец) 
'''
@catch_errors()
def preprocessing_file_excel(path_file_excel: str):
    file_excel =Path(path_file_excel).name
    workbook = None
    try:
        workbook = openpyxl.load_workbook(path_file_excel)
    except Exception as e:
        terminate_script(f'''{file_excel}: Ошибка обработки файла. Возможно открыт обрабатываемый файл. Закройте этот файл и снова запустите скрипт.
                              Ошибка: {e}''')
    sheet = workbook.active

    # Снимаем объединение ячеек
    merged_cells_ranges = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_cells_ranges:
        sheet.unmerge_cells(str(merged_cell_range))

    # Столбец с уровнями группировок
    sheet.insert_cols(idx=1)
    for row_index in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=1)
        cell.value = sheet.row_dimensions[row_index].outline_level
    sheet['A1'] = "Уровень"

    file_excel_treatment = f'{config.folder_path_preprocessing}/preprocessing_{file_excel}'
    if not os.path.exists(f'{config.folder_path_preprocessing}'):
        os.makedirs(f'{config.folder_path_preprocessing}')
    workbook.save(file_excel_treatment)
    workbook.close()
    return file_excel_treatment
